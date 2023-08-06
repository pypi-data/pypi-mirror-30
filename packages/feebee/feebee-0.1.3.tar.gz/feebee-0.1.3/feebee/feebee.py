import os
import sys
import sqlite3
import warnings
import numpy as np
import csv
import pandas as pd
import shutil
import locale
import psutil
from graphviz import Digraph
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
from sas7bdat import SAS7BDAT
from contextlib import contextmanager
from itertools import groupby, chain, zip_longest, accumulate, repeat
from openpyxl import load_workbook
import random
import string
from datetime import datetime
from dateutil.relativedelta import relativedelta
from functools import wraps

warnings.filterwarnings('ignore')
import statsmodels.api as sm


if os.name == 'nt':
    locale.setlocale(locale.LC_ALL, 'English_United States.1252')
else:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')


WORKSPACE = ''
_EMPTY = ''
_filename, _ = os.path.splitext(os.path.basename(sys.argv[0]))
_DBNAME = _filename + '.db'
_GRAPH_NAME = _filename + '.gv'


@contextmanager
def _connect(dbfile, cache_size=100000, temp_store=2):
    conn = _Connection(dbfile, cache_size, temp_store)
    try:
        yield conn
    finally:
        # should I close the cursor?
        conn._cursor.close()
        conn._conn.commit()
        conn._conn.close()


class _Connection:
    def __init__(self, dbfile, cache_size, temp_store):
        global WORKSPACE
        if not WORKSPACE:
            WORKSPACE = os.getcwd()

        dbfile = os.path.join(WORKSPACE, dbfile)

        self._conn = sqlite3.connect(dbfile)
        self._conn.row_factory = _dict_factory 

        self._cursor = self._conn.cursor()
        self._cursor.execute(f'PRAGMA cache_size={cache_size}')
        self._cursor.execute('PRAGMA count_changes=0')
        self._cursor.execute(f'PRAGMA temp_store={temp_store}')
        self._cursor.execute('PRAGMA journal_mode=OFF')

    def fetch(self, tname, where=None, by=None):
        query = f'select * from {tname}'
        if by and by.strip() != '*':
            query += " order by " + by 

        rows = self._conn.cursor().execute(query) 

        if where:
            rows = (r for r in rows if where(r))

        if by:
            gby = groupby(rows, _build_keyfn(by))
            yield from (list(rs) for _, rs in gby)
        else:
            yield from rows

    def insert(self, rs, name):
        self.drop(name)
        rs = iter(rs)

        try:
            r0, rs = _peek_first(rs)
        except StopIteration:
            print(f'No Rows to Create: {name}')
            return
        cols = list(r0)
        n = len(cols)
        self._cursor.execute(_create_statement(name, cols))
        istmt = _insert_statement(name, n)
        self._cursor.executemany(istmt, (list(r.values()) for r in rs if isinstance(r, dict)))

    def load(self, filename, name=None, encoding='utf-8', fn=None):
        if isinstance(filename, str):
            fname, ext = os.path.splitext(filename)

            if ext == '.csv':
                seq = _read_csv(filename, encoding)
            elif ext == '.xlsx':
                seq = _read_excel(filename)
            elif ext == '.sas7bdat':
                seq = _read_sas(filename)
            else:
                raise ValueError('Unknown file extension', ext)
            name = name or fname
        else:
            seq = filename

        if name in self.get_tables():
            return
        if fn:
            seq = (fn(r) for r in seq)

        self.insert(seq, name)

    def tocsv(self, tname, where=None, encoding='utf-8'):
        seq = self.fetch(tname, where=where)
        r0, rs = _peek_first(seq)
        columns = list(r0.keys())
        filename = tname + '.csv'
        with open(os.path.join(WORKSPACE, filename), 'w', newline='',
                  encoding=encoding) as f:
            w = csv.writer(f, delimiter=',')
            w.writerow(columns)
            for r in rs:
                w.writerow(r.values())

    def get_tables(self):
        query = self._cursor.execute("select * from sqlite_master where type='table'")
        return [row['name'] for row in query]

    def drop(self, tables):
        tables = _listify(tables)
        for table in tables:
            self._cursor.execute(f'drop table if exists {table}')

    # may or may not be deprecated
    def rename(self, old, new):
        if old in self.get_tables():
            self._cursor.execute(f'drop table if exists { new }')
            self._cursor.execute(f'alter table { old } rename to { new }')

    # name must be specified
    def join(self, tinfos, name=None):
        tname0, _, mcols0 = tinfos[0]
        join_clauses = []
        for i, (tname1, _, mcols1) in enumerate(tinfos[1:], 1):
            eqs = []
            for c0, c1 in zip(_listify(mcols0), _listify(mcols1)):
                if c1:
                    eqs.append(f't0.{c0} = t{i}.{c1}')
            join_clauses.append(f"left join {tname1} as t{i} on {' and '.join(eqs)}")
        jcs = ' '.join(join_clauses)

        allcols = []
        for i, (_, cols, _) in enumerate(tinfos):
            for c in _listify(cols):
                if c == '*':
                    allcols += [f't{i}.{c1}' for c1 in self._cols(f'select * from {tinfos[i][0]}')]
                else:
                    allcols.append(f't{i}.{c}')

        query = f"create table {name} as select {', '.join(allcols)} from {tname0} as t0 {jcs}"
        self._cursor.execute(query)

    def pwork(self, fn, tname, args):
        n = len(args)
        rndstr = _random_string()
        tempdbs = ['temp' + rndstr + str(i) for i in range(n)]
        try:
            with _connect(tempdbs[0]) as c:
                c.insert(self.fetch(tname), tname)
            for tempdb in tempdbs[1:]:
                shutil.copyfile(os.path.join(WORKSPACE, tempdbs[0]),
                                os.path.join(WORKSPACE, tempdb))

            with ProcessPoolExecutor(max_workers=psutil.cpu_count(logical=False)) as exe:
                exe.map(fn, tempdbs, args)

            with _connect(tempdbs[0]) as c:
                tables = [t for t in c.get_tables() if t != tname]

            for table in tables:
                self._collect(table, tempdbs)
        finally:
            for tempdb in tempdbs:
                fname = os.path.join(WORKSPACE, tempdb)
                if os.path.isfile(fname):
                    os.remove(fname)

    def _collect(self, tname, dbnames):
        with _connect(dbnames[0]) as c:
            cols = c._cols(f"select * from {tname}")

        self.drop(tname)
        self._cursor.execute(_create_statement(tname, cols))
        ismt = _insert_statement(tname, len(cols))
        for dbname in dbnames:
            with _connect(dbname) as c:
                self._cursor.executemany(ismt, (list(r.values()) for r in c.fetch(tname)))

    def _cols(self, query):
        return [c[0] for c in self._cursor.execute(query).description]


def _dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


def _add(**kwargs):
    def fn(r):
        for k, v in kwargs.items():
            r[k] = v(r)
        return r
    return fn


def genfn(c, fn, input, where, by, arg=None):
    if arg:
        for rs in c.fetch(input, where, by):
            val = fn(rs, arg)
            if isinstance(val, dict):
                yield val
            else:
                yield from val
    else:
        if isinstance(fn, dict):
            fn = _add(**fn)
        for rs in c.fetch(input, where, by):
            val = fn(rs)
            if isinstance(val, dict):
                yield val
            else:
                yield from val


def buildfn(dbfile, argset):
    fn, input, where, by, arg, output = argset
    with _connect(dbfile) as c:
        c.insert(genfn(c, fn, input, where, by, arg), output)


def _run(c, job):
    cmd = job['cmd']
    if cmd == 'load':
        fn = _add(**job['fn']) if isinstance(job['fn'], dict) else job['fn']
        c.load(job['file'], job['output'], job['encoding'], fn)
    elif cmd == 'map':
        seq = genfn(c, job['fn'], job['inputs'][0], job['where'], job['by'])
        c.insert(seq, job['output'])
    elif cmd == 'parallel':
        c.pwork(buildfn, job['inputs'][0],
                list(zip(repeat(job['fn']),
                         repeat(job['inputs'][0]),
                         repeat(job['where']),
                         repeat(job['by']),
                         job['args'],
                         repeat(job['output']))))
    elif cmd == 'join':
        c.join(job['args'], job['output'])
    elif cmd == 'union':
        def gen():
            for input in job['inputs']:
                for r in c.fetch(input):
                    yield r
        c.insert(gen(), job['output'])


def load(file=None, fn=None, encoding='utf-8'):
    return {'cmd': 'load',
            'file': file, 
            'fn': fn, 
            'encoding': encoding,
            'inputs': []}


def map(fn=None, data=None, where=None, by=None, args=None):
    result = {
        'cmd': 'map',
        'fn': fn,
        'inputs': [data],
        'where': where,
        'by': by
    }
    if args:
        result['cmd'] = 'parallel'
        result['args'] = args
    return result 


def join(*args):
    inputs = [arg[0] for arg in args]
    return {
        'cmd': 'join',
        'inputs': inputs,
        'args': args
    }


def union(*args):
    inputs = [arg[0] for arg in args]
    return {
        'cmd': 'union',
        'inputs': inputs,
        'args': args
    }


def process(**kwargs):
    def append_output(kwargs):
        for k, v in kwargs.items():
            v['output'] = k
        return [v for _, v in kwargs.items()]
    
    def find_required_tables(jobs):
        tables = set()
        for job in jobs:
            for table in job['inputs']:
                tables.add(table)
            tables.add(job['output'])
        return tables

    # depth first search
    def dfs(data, path, paths=[]):
        datum = path[-1]
        if datum in data:
            for val in data[datum]:
                new_path = path + [val]
                paths = dfs(data, new_path, paths)
        else:
            paths += [path]
        return paths

    def build_graph(jobs):
        graph = {}
        for job in jobs:
            for ip in job['inputs']:
                if graph.get(ip):
                    graph[ip].add(job['output'])
                else:
                    graph[ip] = {job['output']}
        for x in graph:
            graph[x] = list(graph[x])
        return graph

    def render_graph(graph, jobs):
        dot = Digraph()
        for k, v in graph.items():
            dot.node(k, k)
            if k != v:
                for v1 in v:
                    dot.edge(k, v1)
        for job in jobs:
            if job['cmd'] == 'load':
                dot.node(job['output'], job['output'])
        dot.render(_GRAPH_NAME)

    jobs = append_output(kwargs)
    required_tables = find_required_tables(jobs)
    with _connect(_DBNAME) as c:
        def delete_after(missing_table, paths):
            for path in paths:
                if missing_table in path:
                    for x in path[path.index(missing_table):]:
                        c.drop(x)

        def get_missing_tables():
            existing_tables = c.get_tables()
            return [table for table in required_tables
                    if table not in existing_tables]

        def find_jobs_to_do(jobs):
            missing_tables = get_missing_tables()
            result = []
            for job in jobs:
                for table in (job['inputs'] + [job['output']]):
                    if table in missing_tables:
                        result.append(job)
                        break
            return result

        def is_doable(job):
            missing_tables = get_missing_tables()
            return all(table not in missing_tables for table in job['inputs']) \
                and job['output'] in missing_tables

        graph = build_graph(jobs)
        try:
            render_graph(graph, jobs)
        except Exception:
            pass

        starting_points = [job['output'] for job in jobs if job['cmd'] == 'load']
        paths = []
        for sp in starting_points:
            paths += dfs(graph, [sp], [])

        for mt in get_missing_tables():
            delete_after(mt, paths)

        jobs_to_do = find_jobs_to_do(jobs)
        print(f"Starting Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f'To Create: {[j["output"] for j in jobs_to_do]}')
        while jobs_to_do:
            cnt = 0
            for i, job in enumerate(jobs_to_do):
                if is_doable(job):
                    _run(c, job)
                    tm = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    if job['output'] in c.get_tables():
                        print(f"Created: {job['output']} at {tm}")
                    else:
                        print(f"Failed: {job['output']} at {tm}")

                    del jobs_to_do[i]
                    cnt += 1
            if cnt == 0:
                print(f'Failed to Create: {[j["output"] for j in jobs_to_do]}')
                break


def dconv(date, infmt, outfmt=None, **size):
    """Date arithmetic
    Returns int if input(date) is int else str
    """
    outfmt = outfmt or infmt
    if not size:
        # Just convert the format
        return datetime.strftime(datetime.strptime(str(date), infmt), outfmt)
    d1 = datetime.strptime(str(date), infmt) + relativedelta(**size)
    d2 = d1.strftime(outfmt)
    return int(d2) if isinstance(date, int) else d2


# If the return value is True it is converted to 1 or 0 in sqlite3
# istext is unncessary for validity check
def isnum(*xs):
    "Tests if x is numeric"
    try:
        for x in xs:
            float(x)
        return True
    except (ValueError, TypeError):
        return False


def grouper(iterable, n, fillvalue=None):
    "Collect data into fixed-length chunks or blocks"
    # grouper('ABCDEFG', 3, 'x') --> ABC DEF Gxx
    args = [iter(iterable)] * n
    return zip_longest(fillvalue=fillvalue, *args)


def _random_string(nchars=20):
    "Generates a random string of lengh 'n' with alphabets and digits. "
    chars = string.ascii_letters + string.digits
    return ''.join(random.SystemRandom().choice(chars)
                   for _ in range(nchars))


def _peek_first(seq):
    """
    Note:
        peeked first item is pushed back to the sequence
    Args:
        seq (Iter[type])
    Returns:
        Tuple(type, Iter[type])
    """
    # never use tee, it'll eat up all of your memory
    seq1 = iter(seq)
    first_item = next(seq1)
    return first_item, chain([first_item], seq1)


# performance doesn't matter for this, most of the time
def _listify(x):
    """
    Example:
        >>> listify('a, b, c')
        ['a', 'b', 'c']

        >>> listify(3)
        [3]

        >>> listify([1, 2])
        [1, 2]
    """
    try:
        return [x1.strip() for x1 in x.split(',')]
    except AttributeError:
        try:
            return list(iter(x))
        except TypeError:
            return [x]


def _build_keyfn(key):
    " if key is a string return a key function "
    # if the key is already a function, just return it
    if hasattr(key, '__call__'):
        return key
    colnames = _listify(key)
    # special case
    if colnames == ['*']:
        return lambda r: 1

    if len(colnames) == 1:
        col = colnames[0]
        return lambda r: r[col]
    else:
        return lambda r: [r[colname] for colname in colnames]


# primary keys are too much for non-experts
def _create_statement(name, colnames):
    """create table if not exists foo (...)

    Note:
        Every type is numeric.
        Table name and column names are all lowercased
    """
    # every col is numeric, this may not be so elegant but simple to handle.
    # If you want to change this, Think again
    schema = ', '.join([col + ' ' + 'numeric' for col in colnames])
    return "create table if not exists %s (%s)" % (name, schema)


def _insert_statement(name, ncol):
    "insert into foo values (?, ?, ?, ...)"
    qmarks = ', '.join(['?'] * ncol)
    return "insert into %s values (%s)" % (name, qmarks)


def _read_csv(filename, encoding='utf-8'):
    def is_empty_line(line):
        return [x for x in line if x.strip() != ""] == []

    with open(os.path.join(WORKSPACE, filename), encoding=encoding) as f:
        first_line = f.readline()[:-1]
        columns = _listify(first_line)
        ncol = len(columns)

        # NULL byte error handling
        reader = csv.reader(x.replace('\0', _EMPTY) for x in f)
        for line_no, line in enumerate(reader, 2):
            if len(line) != ncol:
                if is_empty_line(line):
                    continue
                raise ValueError(f"Invalid Line at {line_no}: {line}")
            yield {k: v for k, v in zip(columns, line)}


def _read_sas(filename):
    filename = os.path.join(WORKSPACE, filename)
    with SAS7BDAT(filename) as f:
        reader = f.readlines()
        # lower case
        header = next(reader)
        for line in reader:
            yield {k: v for k, v in zip(header, line)}


# this could be more complex but should it be?
def _read_excel(filename):
    def read_df(df):
        cols = df.columns
        for _, r in df.iterrows():
            yield {k: v for k, v in zip(cols, ((str(r[c]) for c in cols)))}

    filename = os.path.join(WORKSPACE, filename)
    # it's OK. Excel files are small
    df = pd.read_excel(filename)
    yield from read_df(df)


def readxl(fname, sheet_name=None, encoding='utf-8'):
    def conv(x):
        try:
            return locale.atoi(x)
        except ValueError:
            try:
                return locale.atof(x)
            except Exception:
                return x

    fname = os.path.join(WORKSPACE, fname)
    if fname.endswith('.csv'):
        with open(fname, encoding=encoding) as fin:
            for rs in csv.reader(x.replace('\0', '') for x in fin):
                yield [conv(x) for x in rs]
    else:
        workbook = load_workbook(fname)
        if not sheet_name:
            sheet_name = workbook.sheetnames[0]
        for row in workbook[sheet_name].iter_rows():
            yield [c.value for c in row]


def drop(tables):
    with _connect(_DBNAME) as c:
        c.drop(tables)


def tocsv(tname, where=None, encoding='utf-8'):
    with _connect(_DBNAME) as c:
        c.tocsv(tname, where, encoding)


def avg(rs, col, wcol=None, ndigits=None):
    if wcol:
        xs = [r for r in rs if isnum(r[col], r[wcol])] 
        val = np.average([x[col] for x in xs], weights=[x[wcol] for x in xs]) 
    else:
        xs = [r for r in rs if isnum(r[col])] 
        val = np.average([x[col] for x in xs])
    return round(val, ndigits) if ndigits else val


def ols(rs, y, *xs):
    df = pd.DataFrame(rs)
    return sm.OLS(df[[y]], sm.add_constant(df[xs])).fit()


def chunk(rs, n, column=None):
    """
    Usage:
        |  chunk(rs, 3) => returns 3 rows about the same size
        |  chunk(rs, [0.3, 0.4, 0.3]) => returns 3 rows of 30%, 40%, 30%
        |  chunk(rs, [100, 500, 1000], 'col')
        |      => returns 4 rows with break points 100, 500, 1000 of 'col'
    """
    size = len(rs)
    if isinstance(n, int):
        start = 0
        result = []
        for i in range(1, n + 1):
            end = int((size * i) / n)
            # must yield anyway
            result.append(rs[start:end])
            start = end
        return result
    # n is a list of percentiles
    elif not column:
        # then it is a list of percentiles for each chunk
        assert sum(n) <= 1, f"Sum of percentils for chunks must be <= 1.0"
        ns = [int(x * size) for x in accumulate(n)]
        result = []
        for a, b in zip([0] + ns, ns):
            result.append(rs[a:b])
        return result
    # n is a list of break points
    else:
        rs.sort(lambda r: r[column]) 
        start, end = 0, 0
        result = []
        for bp in n:
            while (rs[end][column] < bp) and end < size:
                end += 1
            result.append(rs[start:end])
            start = end
        result.append(rs[end:])
        return result
