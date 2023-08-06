from __future__ import print_function

from synthlib.data import Table
from synthlib.utils import p1

from openpyxl import load_workbook


class ExcelTable(Table):
    """A rectangular table range taken from an Excel-table.

        :param workbook: the filename of the Excel file (workbook)
        :param worksheet: the work sheet in the Excel file
        :param row_min: the index of the first row (1-index)
        :param row_max: the index of the last row
        :param col_min: the index of the first column (1-index)
        :param col_max: the index of the last column
    """

    def __init__(self, workbook, worksheet):
        self._workbook_name = workbook
        self._worksheet_name = worksheet

        self._workbook = load_workbook(workbook, read_only=True)
        self._worksheet = self._workbook[str(worksheet)]
        super(ExcelTable, self).__init__()

    def cell(self, r, c):
        """Retrieve a given cell.

        :param row:
        :param column:
        :return:
        """
        return self._worksheet.cell(row=r+1, column=c+1).value

    def row(self, r, c_min=None, c_max=None):
        for rr in self._worksheet.iter_rows(min_col=p1(c_min), max_col=p1(c_max), min_row=p1(r), max_row=p1(r)):
            for cc in rr:
                yield cc.value

    def column(self, c, r_min=None, r_max=None):
        for rr in self._worksheet.iter_rows(min_col=p1(c), max_col=p1(c), min_row=p1(r_min), max_row=p1(r_max)):
            for cc in rr:
                yield cc.value

    @property
    def row_count(self):
        return self._worksheet.max_row

    @property
    def column_count(self):
        return self._worksheet.max_column


