import traceback
import os
import logging
import warnings

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

from pprint import pprint

from ..client import get_client

module_path = os.path.realpath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class Unitcleaner:
    def __init__(self, **data):
        self.data = data


class CleanerConfigurator:
    def __init__(self, cleaner_uid):
        """
        Parameters
        ----------
        cleaner_uid: cleaner_id or (project_id, cleaner_name)
        client: optional
        """

        self.client = get_client()

        if isinstance(cleaner_uid, str):
            self.cleaner_id = cleaner_uid
        else:
            project_id, cleaner_name = cleaner_uid
            cleaners_l = self.client.list(
                "odata/cleaners/",
                params=dict(project=project_id, name=cleaner_name))["data"]
            assert len(cleaners_l) == 1, "One and only one cleaner should exist, found %s. " \
                                         "(project id: %s, cleaner_name: %s)" % (len(cleaners_l),
                                                                                 project_id, cleaner_name)
            self.cleaner_id = cleaners_l[0]["id"]

    # platform to excel
    def iter_unitcleaners(self):
        return self.client.list_iter_all(
            "odata/unitcleaners/",
            params=dict(cleaner=self.cleaner_id)
        )

    def get_cleaner_data(self):
        return self.client.retrieve(
            "odata/cleaners/",
            self.cleaner_id
        )

    def get_project_data(self):
        return self.client.retrieve(
            "odata/projects/",
            self.get_cleaner_data()['project']
        )

    def iter_importer_series(self):
        return self.client.list_iter_all(
            "odata/importer_series",
            params=dict(generator=self.get_cleaner_data()['related_importer'])
        )

    def get_non_configured_series(self):
        not_configured_series = []
        for s in self.iter_importer_series():
            if s['unitcleaner'] is None:
                not_configured_series.append(s)
        return not_configured_series

    def platform_to_excel(self, xlsx_generation_path):
        if xlsx_generation_path[-5:] != '.xlsx':
            logger.warning("The path you have chosen doesn't end up with the extension" + ' ".xlsx ." '
                           + 'You may encounter some issues to open it.')

        project_data = self.get_project_data()
        cleaner_data = self.get_cleaner_data()
        configured_series = list(self.iter_unitcleaners())
        not_configured_series = self.get_non_configured_series()

        for k in ("id", "input_series", "last_clear", "last_run", "cleaner"):
            for d in configured_series:
                d.pop(k, None)

        configured_nb = len(configured_series)
        non_configured_nb = len(not_configured_series)
        tot_nb = configured_nb + non_configured_nb
        start_row = 8
        last_row = tot_nb + start_row - 1
        start_col = 1
        last_col = 23

        # *******************
        # Load model workbook
        # *******************

        wb = load_workbook(os.path.join(module_path, "resources", 'multiclean_config_model.xlsx'))
        ws = wb["Series"]

        ws['B2'].value = project_data['name']
        ws['B3'].value = cleaner_data['name']

        # ****************
        # Validation Lists
        # ****************

        validation = dict(
            save_config=dict(
                dv=DataValidation(type="list", formula1='"x"', allow_blank=True),
                col="C"
            ),
            active=dict(
                dv=DataValidation(type="list", formula1='"yes, no"', allow_blank=True),
                col="D"
            ),
            input_ext=dict(
                dv=DataValidation(type="list", formula1='"instantaneous, delta, mean, bonjour"', allow_blank=True),
                col="F"
            ),
            convention=dict(
                dv=DataValidation(type="list", formula1='"left, right"', allow_blank=True),
                col="G"
            ),
            clock=dict(
                dv=DataValidation(type="list", formula1='"dst, gmt, tzt"', allow_blank=True),
                col="H"
            ),
            timezone=dict(
                dv=DataValidation(type="list", formula1='"Europe/Paris, Africa/Abidjan, Africa/Accra"',
                                  allow_blank=True),
                col="I"
            ),
            input_is_regular=dict(
                dv=DataValidation(type="list", formula1='"yes, no"',
                                  allow_blank=True),
                col="L"
            ),
            output_ext=dict(
                dv=DataValidation(type="list", formula1='"instantaneous, delta, mean, mean_derivate"',
                                  allow_blank=True),
                col="M"
            ),
            reasmple=dict(
                dv=DataValidation(type="list", formula1='"mean, sum, first"', allow_blank=True),
                col="N"
            )
        )

        for name, d in validation.items():
            d["dv"].error = "Your entry is not in the list"
            d["dv"].errorTitle = 'Invalid Entry'
            ws.add_data_validation(d["dv"])
            for row in range(start_row, last_row + 1):
                d["dv"].add(ws['{0}{1}'.format(d["col"], row)])

        # ***************
        # Filling Columns
        # ***************

        # Adding configured series and their parameters
        for i in range(configured_nb):
            row = i + start_row
            ws['A{}'.format(row)].value = configured_series[i]['external_name']
            ws['B{}'.format(row)].value = configured_series[i]['name']
            # Column C = Save Config, always empty when the Excel is generated
            ws['D{}'.format(row)].value = 'yes' if configured_series[i]['active'] else 'no'
            ws['E{}'.format(row)].value = configured_series[i]['freq']
            ws['F{}'.format(row)].value = configured_series[i]['input_unit_type']
            ws['G{}'.format(row)].value = configured_series[i]['input_convention']
            ws['H{}'.format(row)].value = configured_series[i]['clock']
            ws['I{}'.format(row)].value = configured_series[i]['timezone']
            ws['J{}'.format(row)].value = configured_series[i]['unit']
            ws['K{}'.format(row)].value = configured_series[i]['label']
            ws['L{}'.format(row)].value = 'yes' if \
                configured_series[i]['input_expected_regular'] else 'no'  # boolean could be better managed.
            ws['M{}'.format(row)].value = configured_series[i]['unit_type']
            ws['N{}'.format(row)].value = configured_series[i]['resample_rule']
            ws['O{}'.format(row)].value = configured_series[i]['interpolate_limit']
            ws['P{}'.format(row)].value = configured_series[i]['wait_offset']
            ws['Q{}'.format(row)].value = configured_series[i]['operation_fct']
            ws['R{}'.format(row)].value = configured_series[i]['filter_fct']
            ws['S{}'.format(row)].value = configured_series[i]['derivative_filter_fct']
            ws['T{}'.format(row)].value = configured_series[i]['custom_delay']
            ws['U{}'.format(row)].value = configured_series[i]['custom_fct']
            ws['V{}'.format(row)].value = configured_series[i]['custom_before_offset']
            ws['W{}'.format(row)].value = configured_series[i]['custom_after_offset']

        # Adding non-configured series external names
        for i in range(non_configured_nb):
            row = i + start_row + configured_nb
            ws['A{}'.format(row)].value = not_configured_series[i]['external_name']

        # *********
        # Styling + Conditional Formatting
        # *********

        greenFill = PatternFill(start_color='66CC99', end_color='66CC99', fill_type='solid')

        center_align = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)

        default_border = Border(
            left=Side(border_style=None, color='FF000000'),
            right=Side(border_style=None, color='FF000000'),
            top=Side(border_style='thin', color='b0b0b0'),
            bottom=Side(border_style='thin', color='b0b0b0')
        )

        right_thin_border = Border(
            right=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='b0b0b0')
        )
        bottom_thin_border = Border(bottom=Side(border_style='thin', color='000000'))
        right_bottom_thin_border = Border(
            bottom=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )

        for row in range(start_row, last_row + 1):
            for col in range(start_col, last_col + 1):
                # Style
                cell = ws.cell(row=row, column=col)
                cell.border = default_border
                cell.alignment = center_align
                if col == 3:
                    cell.font = Font(bold=True)
                if col != last_col and row == last_row:
                    cell.border = bottom_thin_border
                elif col == last_col and row != last_row:
                    cell.border = right_thin_border
                elif col == last_col and row == last_row:
                    cell.border = right_bottom_thin_border
                # Formatting
                ws.conditional_formatting.add(
                    '{0}{1}'.format(cell.column, cell.row),
                    FormulaRule(formula=['C{}="x"'.format(row)],
                                fill=greenFill)
                )

        # **********
        # write book
        # **********
        wb.save(xlsx_generation_path)
        print('The file "multiclean_config.xlsx" has been successfully generated at ' + xlsx_generation_path)

    # excel to platform
    def configure_unit_cleaner(self, data, update_if_exists=False):
        unitcleaners = self.client.list_iter_all(
            "odata/unitcleaners/",
            params=dict(cleaner=self.cleaner_id, external_name=data["external_name"])
        )
        unitcleaner_l = list(unitcleaners)
        if len(unitcleaner_l) == 1:
            if not update_if_exists:
                raise AssertionError("unitcleaner already exists, can't create (use update_if_exists=True)")
            # we delete
            self.client.destroy("odata/unitcleaners/", unitcleaner_l[0]["id"])

        # add cleaner argument to data (will replace if already there)
        data["cleaner"] = self.cleaner_id

        # we create
        self.client.create("odata/unitcleaners/", data)
        print(data['name'] + ' has been successfully configured')

    def excel_to_platform(self, xlsx_path, update_if_exists=False):
        unitcleaners = batch_configure(xlsx_path)
        for uc in unitcleaners:
            try:
                data = dict(cleaner=self.cleaner_id)
                data.update(uc.data)
                self.configure_unit_cleaner(data, update_if_exists=update_if_exists)
            except Exception:
                name = getattr(uc, "external_name", "Unknown (no external name provided).")
                print("Error while configuring unitcleaner: %s\n%s\n\n" % (name, traceback.format_exc()))


def batch_configure(path, max_input_length=20000):
    unitcleaners = []
    wb = load_workbook(path)
    ws = wb["Series"]
    start_row = 8
    for row in range(start_row, max_input_length + start_row):
        # print(ws['C{}'.format(row)].value)
        if ws['C{}'.format(row)].value == 'x':
            if ws['D{}'.format(row)].value == 'yes':
                ws['D{}'.format(row)].value = 'true'
            elif ws['D{}'.format(row)].value == 'no':
                ws['D{}'.format(row)].value = 'false'
            unitcleaners.append(Unitcleaner(
                # An empty cell is None
                external_name=ws['A{}'.format(row)].value,
                name=ws['B{}'.format(row)].value if ws['B{}'.format(row)].value else ws['A{}'.format(row)].value,
                active=ws['D{}'.format(row)].value,
                freq=ws['E{}'.format(row)].value,
                input_unit_type=ws['F{}'.format(row)].value,
                input_convention=ws['G{}'.format(row)].value,
                clock=ws['H{}'.format(row)].value,
                timezone=ws['I{}'.format(row)].value,
                unit=ws['J{}'.format(row)].value if ws['J{}'.format(row)].value is not None else '',
                label=ws['K{}'.format(row)].value if ws['K{}'.format(row)].value is not None else '',
                input_expected_regular= ws['L{}'.format(row)].value == 'yes', #Boolean could be better managed.
                unit_type=ws['M{}'.format(row)].value if ws['M{}'.format(row)].value is not None else ws['F{}'.format(row)].value,
                resample_rule=ws['N{}'.format(row)].value if ws['N{}'.format(row)].value is not None else 'mean',
                interpolate_limit=ws['O{}'.format(row)].value if ws['O{}'.format(row)].value is not None else 0,
                wait_offset=ws['P{}'.format(row)].value if ws['P{}'.format(row)].value is not None else '6H',
                operation_fct=ws['Q{}'.format(row)].value,
                filter_fct=ws['R{}'.format(row)].value,
                derivative_filter_fct=ws['S{}'.format(row)].value,
                custom_delay=ws['T{}'.format(row)].value,
                custom_fct=ws['U{}'.format(row)].value,
                custom_before_offset=ws['V{}'.format(row)].value,
                custom_after_offset=ws['W{}'.format(row)].value
            ))

    return unitcleaners
