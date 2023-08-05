
import json
import os
import re
import subprocess
import sys
import textwrap

try:
	import configparser # python 3.x name
except ImportError:
	import ConfigParser as configparser # python 2.x name

import click
import openpyxl



# -------------------------------------------------------------------
def get_config(config_file_path=None):
	cp = configparser.SafeConfigParser()

	if config_file_path:
		cp.read(config_file_path)
	else:
		cp.add_section('layout')
		cp.set('layout', 'HeaderRow', '1')
		cp.set('layout', 'FirstQuestionRow', '2')
		cp.add_section('columnNames')
		cp.set('columnNames', 'QId',      'Question Code')
		cp.set('columnNames', 'QText',    'Question / Statement')
		cp.set('columnNames', 'RespList', 'Possible responses')
		cp.set('columnNames', 'Resp1',    'If they choose response 1')
		cp.set('columnNames', 'Resp2',    'If they choose response 2')
		cp.set('columnNames', 'Resp3',    'If they choose response 3')
		cp.set('columnNames', 'Resp4',    'If they choose response 4')
		cp.set('columnNames', 'Resp5',    'If they choose response 5')
		cp.set('columnNames', 'Resp6',    'If they choose response 6')
		cp.set('columnNames', 'Resp7',    'If they choose response 7')
		cp.set('columnNames', 'Resp8',    'If they choose response 8')
		cp.set('columnNames', 'Resp9',    'If they choose response 9')

	return cp

# -------------------------------------------------------------------
def open_excel_sheet(excel_file_path, sheet_name=None):
	wb = openpyxl.load_workbook(filename=excel_file_path, read_only=True)
	sheet_count = len(wb.sheetnames)

	if sheet_count < 1:
		sys.exit('no sheets in excel file provided.')

	if sheet_name == None:
		sheet_name = wb.sheetnames[0]

	if sheet_name in wb.sheetnames:
		return wb[sheet_name]
	else:
		sys.exit('Worksheet named "' + sheet_name + '" not found in Workbook.')

# -------------------------------------------------------------------
def identify_column_indexes(sheet, cfg):
	''' given a sheet and a config, return a data structure with 
	canonical column names as keys, and the index into a row for
	that column (indexed from 0) as values.
	'''
	cols = {}

	headerrow = cfg.getint('layout', 'headerrow')
	headers = sheet.iter_rows(min_row=headerrow, max_row=headerrow)
	# there's just one row, we're for-ing it coz its in a generator
	for r in headers:
		for c in r: # for each cell in this header row
			for col in map(lambda item: item[0], cfg.items('columnNames')):
				if c.value == cfg.get('columnNames', col):
					cols[col] = c.column - 1

	# print(cols)
	return cols

# -------------------------------------------------------------------
def extract_qns_from_sheet(sheet, cfg):
	'''Pulls question rows from a spreadsheet. 
	Returns an array of qn arrays
	'''
	questions = []

	# find useful col indexes
	cols = identify_column_indexes(sheet, cfg)

	# get an iterator for the question-containing rows
	first_row = cfg.getint('layout', 'firstquestionrow')
	row_iterator = sheet.iter_rows(min_row=first_row)

	# loop thru qn rows
	for index, qr in enumerate(row_iterator):
		# get all the question's info
		qid = qr[cols['qid']].value

		qtext = qr[cols['qtext']].value
		qtext = qtext and qtext.strip() or qtext

		resplist = qr[cols['resplist']].value
		resplist = resplist and resplist.strip() or resplist

		# make a list of the response texts
		responses = []
		for r in range(1,10):
			col_index = cols.get('resp' + str(r), None)
			if col_index and qr[col_index].value:
				responses.append(qr[col_index].value.strip())
			else:
				responses.append(None)
		
		if qid and qtext:
			# print(qid, qtext, resplist, responses)
			questions.append([qid, qtext, resplist, responses])

	return questions

# -------------------------------------------------------------------
def qns_to_conversation(qns, default_response="okay"):
	'''converts an array of questions into an ordered conversation.'''
	conversation = {
		'start': qns[0][0], # qid of the first question
		'end': None,		# this gets filled out later
		'questions': {}		# this gets added to
	}

	for index, q in enumerate(qns):
		# is this the last question?
		last = index == len(qns) - 1

		# the question id is the first thing in the qn
		qid = q[0]

		if last:
			conversation['end'] = qid

		# the question's text is the second item
		qtext = q[1]

		# assume the "next" question is the one immediately following
		if last:
			qnext = None
		else:
			qnext = qns[index + 1][0]

		# if there are responses, they're in q[2]
		collect = None
		if q[2]:
			# q[2] could be either a list of responses, or
			# a text response.
			textreq = re.search(r'^\s*\{\s*text\s*\}\s*$', q[2], flags=re.IGNORECASE)

			if textreq:
				# ..it's a request for text input.
				resp_list = []
				collect = 'text'
			else:
				# q[2] isn't blank, and isn't a text req, so it's a proper
				# list of responses.
				#
				# If the possible responses includes a semicolon, use that as
				# the separator. Otherwise use comma
				split_char = ';' if ';' in q[2] else ','
				resp_list = [resp.strip() for resp in q[2].split(split_char)]

		else:
			# if q[2] is blank, assume a default response
			resp_list = [default_response]

		answers = []
		for a_index, a in enumerate(resp_list):

			ans_next = qnext # default to the following qn

			ans_text = q[3][a_index]

			if ans_text and 'GOTO:' in ans_text:
				text_bits = ans_text.split('GOTO:')
				ans_text = text_bits[0].strip()
				if len(ans_text) == 0:
					ans_text = None
				ans_next = text_bits[-1].strip()

			answers.append({
				"id": qid + '::' + str(a_index),
				"label": a,
				"info": ans_text,
				"next": ans_next
			})

		conversation['questions'][qid] = {
			"text": qtext,
			"next": qnext,
			"collect": collect,
			"answers": answers
		}



	return conversation

# -------------------------------------------------------------------
def esc(str):
	''' escape a string for use in a graphviz node. '''
	return (str.replace('{','\{').replace('}','\}').replace('"','\\"'))
# -------------------------------------------------------------------
def prep_string(str, width=65):
	''' prepare a string for use as the label of a graphviz graph node.
	This means wrapping the text using textwrap.wrap, and re-joining the
	lines with a \\n (coz GraphViz uses literal "\n" for newlines).
	That bit splits into lines and re-wraps them with \\n: 
		'\\n'.join(textwrap.wrap(p))
	However we also want && to be on lines of their own (&& is a paragraph
	separator for the converstion displayer). So we split into paragraphs
	using split, wrap each paragraph as above, then re-join paragraphs 
	using \\n&&\\n.
	'''
	paragraphs = map(lambda p: '\\n'.join(textwrap.wrap(esc(p), width=width)), str.split('&&'))
	return '\\n&&\\n'.join(paragraphs)

# -------------------------------------------------------------------
def conversation_to_dot(conversation, background='transparent'):

	dot = []
	dot.append('digraph {')
	dot.append('\tranksep=0.5')
	# dot.append('\tbgcolor="#cceeff"')
	dot.append('\tbgcolor="{}"'.format(background))
	dot.append('\tnode [style="filled",fillcolor="#ffffff",shape="box",fontname="sans",width="1.2"]')

	# start and end nodes..
	dot.append(u'\t"::START::" [label="START",shape="circle",style="filled,bold",width="0.75"]')
	dot.append(u'\t"::END::" [label="END",shape="circle",style="filled,bold",width="0.75"]')

	# link ::START:: to the first node (linking to ::END:: happens for each question)
	dot.append(u'\t"::START::" -> "{}"'.format(conversation['start']))
	dot.append(u'\n')

	for qid, qn in conversation['questions'].items():

		# the node itself
		# \u2014 is an em-dash
		# \u2015 is a "horizontal bar" character similar to an em-dash
		# \u2500 is a horizontal box-drawing character
		dot.append(u'\t"{}" [label="{{\\N}}|{{\u2014\u2014\u2014\\n{}\\n\u2014\u2014\u2014\\n}}",shape="record",style="filled,rounded"]'.format(qid, prep_string(qn['text'])))

		# the node's answers
		dot.append(u'\t{ rank=same; ')
		for ans in qn['answers']:
			dot.append(u'\t\t"{}" [label="{}", shape="box", style="filled"]'.format(esc(ans['id']), esc(ans['label'])))
		dot.append(u'\t}')

		if qn['collect']:
			# if it's a collect qn, draw a text-input node
			if qn['next']:
				next_node = qn['next']
			else:
				next_node = '::END::'

			dot.append(u'\t\t"{}" [label="user enters: {}", shape="parallelogram", style="filled"]'.format(qid + '::collect', '\{' + qid + '\}'))
			dot.append(u'\t\t"{}" -> "{}" -> "{}"'.format(qid, qid + '::collect', next_node))

		# the node's answers' info blocks (which is a zero-length
		# list if this is a text input node)
		for ans in qn['answers']:
			if ans['info']:
				dot.append(u'\t"{}" [label="{}", shape="box", style="filled,rounded"]'.format(esc(ans['id']) + '::info', prep_string(ans['info'], 45)))

		# link the node to its answers and their info blocks

		for ans in qn['answers']:
			if ans['next']:
				next_node = ans['next']
			else:
				next_node = '::END::'

			if ans['info']:
				dot.append(u'\t"{}" -> "{}" -> "{}" -> "{}"'.format(qid, ans['id'], ans['id'] + '::info', next_node))
			else:
				dot.append(u'\t"{}" -> "{}" -> "{}"'.format(qid, ans['id'], next_node))

		dot.append(u'\n')

	dot.append(u'}')
	return u'\n'.join(dot)

# -------------------------------------------------------------------
def save_as_json(conversation, output_file_path):
	'''given a set of questions properly linked, save those as 
	json data'''
	with open(output_file_path, 'w') as out_file:
		json.dump(conversation, out_file, indent=4)

# -------------------------------------------------------------------
def save_as_dot(conversation, output_file_path, bg):
	'''given a set of questions properly linked, draw the 
	conversational flow as a DOT file'''
	with open(output_file_path, 'w') as out_file:
		out_file.write(conversation_to_dot(conversation, bg).encode('utf8'))

# -------------------------------------------------------------------
def good(msg):
	click.echo(click.style(u'\u2714 ', fg='green') + msg)

# -------------------------------------------------------------------
def bad(msg):
	click.echo(click.style(u'\u2718 ', fg='red') + msg)


# ===================================================================
@click.command()
@click.argument('excelfile', type=click.Path(exists=True, dir_okay=False))
@click.argument('questionfile', type=click.Path(writable=True, dir_okay=False), default='conversation.json')
@click.option('--sheet', default=None, help='name of the worksheet (default: workbook\'s first sheet)')
@click.option('--config', default=None, help='optional config file')
def excel2qns(excelfile, questionfile, sheet, config):
	'''Generates a question data file from an Excel document.

	You must specify the input -- an Excel workbook file -- on the
	command line.  The output file defaults to conversation.json but 
	may be specified as the second argument.

	By default the first sheet in your workbook will be read; you
	can specify a different sheet by name with --sheet.

	This tool uses column titles to identify data, and can use a 
	config file to change the row and column assumptions made.
	Use the --config option to specify a config file.

	Run sat5pconfig --help for more infomation on config files.	
	'''
	cfg = get_config(config)
	xlws = open_excel_sheet(excelfile, sheet)

	qns = extract_qns_from_sheet(xlws, cfg)
	convo = qns_to_conversation(qns)

	save_as_json(convo, questionfile)

# ===================================================================
@click.command()
@click.argument('configfile', type=click.Path(writable=True, dir_okay=False), default='config.cfg')
def sat5pconfig(configfile):
	'''Create a config file for use by other sat5ptools commands.

	Provide a configfile path for the new config file. You can then
	edit the config file, and pass it to other sat5ptools commands
	using their --config configfile options.
	'''
	cfg = get_config()
	if os.path.exists(configfile):
		bad('Not writing config to "{}", file already exists.'.format(configfile))
	else:
		with open(configfile, 'w') as cf:
			cfg.write(cf)

		good('Wrote config to file "{}".'.format(configfile))

# ===================================================================
@click.command()
@click.argument('excelfile', type=click.Path(exists=True, dir_okay=False))
@click.argument('dotfile', type=click.Path(writable=True, dir_okay=False), default='conversation.dot')
@click.option('--sheet', default=None, help='name of the worksheet (default: workbook\'s first sheet)')
@click.option('--make', is_flag=True, help='run Graphviz to render the graph')
@click.option('--makeformat', default='png', help='specifies format with rendering with --make')
@click.option('--bg', default='#cceeff', help='background colour (defaults to #cceeff, a light blue)')
@click.option('--gv', default='dot', help='path to Graphviz\'s dot command')
@click.option('--config', help='config file (default: ./satconfig.cfg)')
def excel2graph(excelfile, dotfile, sheet, config, make, makeformat, bg, gv):
	'''Generates a conversation graph from an Excel document.

	You must specify the input -- an Excel workbook file -- on the
	command line.  The output file defaults to conversation.dot but 
	may be specified as the second argument.

	The graph produced is a Graphviz "dot" file, which is only
	useful if you have Graphviz (graphviz.org) installed.  You can
	run Graphviz yourself with a command like this:

	dot -O -Tpng conversation.dot

	..which will create a PNG of your graph, or you can supply the 
	--make option and this tool will attempt to run Graphviz's dot 
	command for you. --make will by default try to make a PNG, but
	you can supply another Graphviz-supported format using
	--makeformat, e.g --make --makeformat svg

	If the dot command isn't in your path, specify the path to the
	Graphviz dot executable with the --gv option. E.g. Windows users 
	might use --gv "c:\Program Files (x86)\Graphviz2.38\bin\dot"

	Specify the graph's background with --bg and a CSS-style six 
	digit hex colour, or the word "transparent" for a transparent 
	background, e.g. --bg #ff8800 for orange. The Graphviz docs have 
	additional info at 
	https://www.graphviz.org/doc/info/attrs.html#d:bgcolor

	By default the first sheet in your workbook will be read; you
	can specify a different sheet by name with --sheet.

	This tool uses a config file to set which row is the "header"
	row (by default, it's the top row), and what the column names
	in your worksheet are. The default config is satconfig.cfg,
	you can use the --config option to use a different config file.
	'''
	cfg = get_config(config)
	xlws = open_excel_sheet(excelfile, sheet)

	qns = extract_qns_from_sheet(xlws, cfg)
	convo = qns_to_conversation(qns)

	save_as_dot(convo, dotfile, bg)

	if make:
		# they want us to try running graphviz
		subprocess.call([gv, '-O', '-T' + makeformat, dotfile])

# ===================================================================
@click.command()
@click.argument('excelfile', type=click.Path(exists=True, dir_okay=False))
@click.argument('outname', default='conversation')
@click.option('--sheet', default=None, help='name of the worksheet (default: workbook\'s first sheet)')
@click.option('--make', is_flag=True, help='run Graphviz to render the graph')
@click.option('--makeformat', default='png', help='specifies format with rendering with --make')
@click.option('--bg', default='#cceeff', help='background colour (defaults to #cceeff, a light blue)')
@click.option('--gv', default='dot', help='path to Graphviz\'s dot command')
@click.option('--config', help='config file (default: ./satconfig.cfg)')
def excel2all(excelfile, outname, sheet, config, make, makeformat, bg, gv):
	'''Given an Excel document, generate both the question file
	and a graph describing that conversation.

	You must specify the input -- an Excel workbook file -- on the
	command line.  The output files will both start with OUTNAME;
	the data file will have a .json extension, and the graph file
	will have .dot added.  OUTNAME defaults to 'conversation'.

	The graph produced is a Graphviz "dot" file, which is only
	useful if you have Graphviz (graphviz.org) installed.  You can
	run Graphviz yourself with a command like this:

	dot -O -Tpng conversation.dot

	..which will create a PNG of your graph, or you can supply the 
	--make option and this tool will attempt to run Graphviz's dot 
	command for you. --make will by default try to make a PNG, but
	you can supply another Graphviz-supported format using
	--makeformat, e.g --make --makeformat svg

	If the dot command isn't in your path, specify the path to the
	Graphviz dot executable with the --gv option. E.g. Windows users 
	might use --gv "c:\Program Files (x86)\Graphviz2.38\bin\dot"

	Specify the graph's background with --bg and a CSS-style six 
	digit hex colour, or the word "transparent" for a transparent 
	background, e.g. --bg #ff8800 for orange. The Graphviz docs have 
	additional info at 
	https://www.graphviz.org/doc/info/attrs.html#d:bgcolor

	By default the first sheet in your workbook will be read; you
	can specify a different sheet by name with --sheet.

	This tool uses a config file to set which row is the "header"
	row (by default, it's the top row), and what the column names
	in your worksheet are. The default config is satconfig.cfg,
	you can use the --config option to use a different config file.
	'''
	cfg = get_config(config)
	xlws = open_excel_sheet(excelfile, sheet)

	qns = extract_qns_from_sheet(xlws, cfg)
	convo = qns_to_conversation(qns)

	questionfile = outname + '.json'
	dotfile = outname + '.dot'

	save_as_json(convo, questionfile)
	save_as_dot(convo, dotfile, bg)

	if make:
		# they want us to try running graphviz
		subprocess.call([gv, '-O', '-T' + makeformat, dotfile])

# ===================================================================

