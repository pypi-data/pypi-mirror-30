"""
Utils for statistical analysis using sqlite3 engine
"""
import os
import sqlite3
import copy as cp
import warnings
import inspect
import numpy as np
import csv
import pandas as pd
import shutil
import collections
import locale 
import psutil 
from graphviz import Digraph
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
from sas7bdat import SAS7BDAT
from contextlib import contextmanager
from itertools import groupby, islice, chain, tee, \
    zip_longest, accumulate, repeat
from openpyxl import load_workbook

from .util import isnum, _listify, _peek_first, _random_string

# pandas raises warnings because maintainers of statsmodels are lazy
warnings.filterwarnings('ignore')
import statsmodels.api as sm

# workspace will be deprecated
WORKSPACE = ''
DBNAME = 'workspace.db'
GRAPH_NAME = 'workspace.gv'

if os.name == 'nt':
    locale.setlocale( locale.LC_ALL, 'English_United States.1252' )
else:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')


@contextmanager
def connect(dbfile, cache_size=100000, temp_store=2):
    # temp_store might be deprecated
    """ Connects to SQL database, decorated with contextmanager

    Usage:
        with connect('dbfile.db') as conn:
            conn.load('sample.csv')

    Args:
        |  dbfile(str): relative path for db file.
        |  cache_size(int):  cache size for insertion.
        |  temp_store(int):  0, 1 for file. 2 for memory.
    """
    splus = SQLPlus(dbfile, cache_size, temp_store)
    try:
        yield splus
    finally:
        # should I close the cursor?
        splus._cursor.close()
        splus.conn.commit()
        splus.conn.close()


# Don't try to be smart, unless you really know well
class Row:
    """Mutable version of sqlite3.row

    Note:
        |  The order of assignment is preserved

        Row value types are one of int, float or str
    """
    # works for python 3.6 and higher
    def __init__(self, **kwargs):
        super().__setattr__('_dict', kwargs)

    @property
    def columns(self):
        """Returns a list of column names(strings)
        """
        return list(self._dict.keys())

    @property
    def values(self):
        """Returns a list of column values
        """
        return list(self._dict.values())

    def copy(self):
        # 
        r0 = cp.copy(self)
        r0.__init__()
        for c, v in zip(self.columns, self.values):
            r0[c] = v
        return r0   
 
    def __getattr__(self, name):
        return self._dict[name]

    def __setattr__(self, name, value):
        self._dict[name] = value

    def __delattr__(self, name):
        del self._dict[name]

    def __getitem__(self, name):
        return self._dict[name]

    def __setitem__(self, name, value):
        self._dict[name] = value

    def __delitem__(self, name):
        del self._dict[name]

    def __repr__(self):
        content = ', '.join(c + '=' + repr(v) for c, v in self._dict.items())
        return 'Row(' + content + ')'

    # for pickling, very important
    def __getstate__(self):
        return self.__dict__

    def __setstate__(self, d):
        self.__dict__.update(d)

    # TODO:
    # hasattr doesn't work properly
    # you can't make it work by changing getters and setters
    # to an ordinary way. but it is slower


class Rows:
    """Wrapper for a list of Row instances

    Attributes:
        rows(list of Row instances)
    """
    # don't try to define __getattr__, __setattr__
    # list objects has a lot of useful attributes that can't be overwritten
    # not the same situation as 'row' class

    # inheriting list can be problematic
    # when you want to use this as a superclass
    # see 'where' method, you must return 'self' but it's not efficient
    # (at least afaik) if you inherit list

    def __init__(self, rows):
        """
        Args:
            rows(sequence of Row instances)
        """
        self.rows = list(rows) if rows else []

    def __len__(self):
        return len(self.rows)

    # __getitem__ enables you to iterate 'Rows'
    def __getitem__(self, k):
        """
        Args:
            k: int, slice, or column name
        """
        if isinstance(k, int):
            return self.rows[k]
        if isinstance(k, slice):
            # shallow copy for non-destructive slicing
            return self._newrows(self.rows[k])

        k = _listify(k)
        if len(k) == 1:
            k0 = k[0]
            return [r[k0] for r in self.rows]
        return [[r[k0] for k0 in k] for r in self.rows]

    # Do not define __setitem__, it will get users confused

    def __delitem__(self, k):
        if isinstance(k, int) or isinstance(k, slice):
            del self.rows[k]
            return

        k = _listify(k)
        if len(k) == 1:
            k0 = k[0]
            for r in self.rows:
                del r[k0]
        else:
            for r in self.rows:
                for k0 in k:
                    del r[k0]

    def __add__(self, other):
        return self._newrows(self.rows + other.rows)

    def _newrows(self, rs):
        # copying rows and build Rows object
        # Am I worring too much?, this is for inheritance
        self.rows, temp = [], self.rows
        other = cp.copy(self)
        other.rows, self.rows = list(rs), temp
        return other

    def copy(self):
        rs0 = []
        for r in self.rows:
            rs0.append(r.copy())
        return self._newrows(rs0)

    # Limited version of __setitem__
    def set(self, k, v):
        for r in self.rows:
            r[k] = v

    def append(self, r):
        self.rows.append(r)
        return self

    # destructive!!!
    def order(self, key, reverse=False):
        """Order rows by key

        Args:
            |  key(str, list of str, fn)
            |  reverse(bool)

        """
        # You can pass fn as key arg but not recommended
        self.rows.sort(key=_build_keyfn(key), reverse=reverse)
        return self

    def where(self, pred):
        """Filters rows

        Args:
            pred(str or fn): "year > 1990 and size < 100" or predicate
        """
        return self._newrows([r for r in self if pred(r)])

    def avg(self, col, wcol=None, ndigits=None):
        """Computes average

        Args:
            |  col(str): column name to compute average
            |  wcol(str): column for weight
            |  n(int): round digits

        Returns float
        """
        if wcol:
            xs = self.where(lambda r: isnum(r[col], r[wcol]))
            val = np.average(xs[col], weights=xs[wcol])
        else:
            xs = self.where(lambda r: isnum(r[col]))
            val = np.average(xs[col])
        return round(val, ndigits) if ndigits else val

    def ols(self, model):
        """OLS fit

        Args:
            model(str): 'col1 ~ col2 + col3'

        Returns:
            |  statsmodels.api.OLS(model).fit()
            |  http://www.statsmodels.org/stable/index.html

        Note:
            Constant is added automatically
        """
        def parse_model(model):
            "y ~ x1 + x2 => ['y', 'x1', 'x2']"
            left, right = model.split('~')
            return [left.strip()] + [x.strip() for x in right.split('+')]

        y, *xs = parse_model(model)
        df = self.df()
        res = sm.OLS(df[[y]], sm.add_constant(df[xs])).fit()
        return res

    def truncate(self, col, limit=0.01):
        """Truncate rows that are out of limits

        Args:
            |  col(str): column name
            |  limit(float): for both sides respectably.

        Returns self
        """
        xs = self[col]
        lower = np.percentile(xs, limit * 100)
        higher = np.percentile(xs, (1 - limit) * 100)
        return self.where(lambda r: r[col] >= lower and r[col] <= higher)

    def winsorize(self, col, limit=0.01):
        """Winsorsize rows that are out of limits

        Args:
            |  col(str): column name.
            |  limit(float): for both sides respectably.

        Returns self
        """
        xs = self[col]
        lower = np.percentile(xs, limit * 100)
        higher = np.percentile(xs, (1 - limit) * 100)
        for r in self.rows:
            if r[col] > higher:
                r[col] = higher
            elif r[col] < lower:
                r[col] = lower
        return self

    # implicit ordering
    def group(self, key):
        """Yields rows of each group

        Args:
            key(str or list of str or fn): columnn name or fn to group

        Returns: list of Rows
        """
        # key can be a fn but not recommended
        keyfn = _build_keyfn(key)
        self.order(keyfn)
        return [self._newrows(list(rs)) for _, rs in groupby(self, keyfn)]

    def overlap(self, size, step=1, group=None):
        if group:
            xs = self.group(group)
            result = []
            for i in range(0, len(xs), step):
                result.append(self._newrows(chain(*xs[i:i + size])))
            return result
        else:
            result = []
            for i in range(0, len(self), step):
                result.append(self[i:i + size])
            return result

    def chunk(self, n, col=None):
        """Yields Rows, useful for building portfolios

        Usage:
            |  self.chunk(3) => yields 3 rows about the same size
            |  self.chunk([0.3, 0.4, 0.3]) => yields 3 rows of 30%, 40%, 30%
            |  self.chunk([100, 500, 1000], 'col')
            |      => yields 4 rows with break points 100, 500, 1000
        """
        size = len(self)
        if isinstance(n, int):
            start = 0
            result = []
            for i in range(1, n + 1):
                end = int((size * i) / n)
                # must yield anyway
                result.append(self[start:end])
                start = end
            return result
        # n is a list of percentiles
        elif not col:
            # then it is a list of percentiles for each chunk
            assert sum(n) <= 1, f"Sum of percentils for chunks must be <= 1.0"
            ns = [int(x * size) for x in accumulate(n)]
            result = []
            for a, b in zip([0] + ns, ns):
                result.append(self[a:b])
            return result
        # n is a list of break points
        else:
            self.order(col)
            start, end = 0, 0
            result = []
            for bp in n:
                while (self[end][col] < bp) and end < size:
                    end += 1
                result.append(self[start:end])
                start = end
            result.append(self[end:])
            return result

    # Use this when you need to see what's inside
    # for example, when you want to see the distribution of data.
    def df(self, cols=None):
        """Returns pandas data frame
        """
        if cols:
            cols = _listify(cols)
            return pd.DataFrame([[r[col] for col in cols] for r in self.rows],
                                columns=cols)
        else:
            cols = self.rows[0].columns
            return pd.DataFrame([r.values for r in self], columns=cols)


class SQLPlus:
    def __init__(self, dbfile, cache_size, temp_store):
        """
        Args:
            |  dbfile (str): db filename or ':memory:'
            |  cache_size(int)
            |  temp_store(int)

        """
        global WORKSPACE

        # set workspace if it's not there
        if not WORKSPACE:
            # default workspace
            WORKSPACE = os.getcwd()

        if dbfile != ':memory:':
            dbfile = os.path.join(WORKSPACE, dbfile)

        # you may want to pass sqlite3.deltypes or something like that
        # but at this moment I think that will make matters worse
        self.conn = sqlite3.connect(dbfile)
        # You can safely uncomment the following line
        self.conn.row_factory = sqlite3.Row

        # default cursor
        self._cursor = self.conn.cursor()
        # some performance tuning
        self._cursor.execute(f'PRAGMA cache_size={cache_size}')

        # Don't be too greedy, comment out the following line.
        # It's too dangerous. Your db file can be corrupted
        # self._cursor.execute('PRAGMA synchronous=OFF')

        self._cursor.execute('PRAGMA count_changes=0')
        # temp store at memory
        self._cursor.execute(f'PRAGMA temp_store={temp_store}')
        self._cursor.execute('PRAGMA journal_mode=OFF')

    def rows(self, tname, cols=None, where=None, order=None):
        """Returns Rows
        """
        return Rows(self.fetch(tname, cols, where, order))

    def df(self, tname, cols=None, where=None, order=None):
        """Returns pandas data frame
        """
        return self.rows(tname, cols, where, order).df(cols)

    def fetch(self, tname, cols=None, where=None,
              order=None, group=None, overlap=None):
        """Generates a sequence of rows from a table.

        Args:
            |  tname(str): table name
            |  cols(str or list of str): columns to fetch
            |  where: predicate function
            |  order(str or list of str): comma separated str
            |  group(str or list of str): comma separated str
            |  overlap: int or (int, int)

        Yields:
            Row or Rows
        """
        def _overlap(seq, size, step):
            """generates chunks of seq for rollover tasks.
            seq is assumed to be ordered
            """

            ss = tee(seq, size)
            # consume
            for i, s in enumerate(ss):
                for i1 in range(i):
                    next(s)
            xss = zip_longest(*ss, fillvalue=None)
            for xs in islice(xss, 0, None, step):
                # lets just go easy
                yield [x for x in xs if x is not None]

        order = _listify(order) if order else []
        group = _listify(group) if group else []

        order = group + order if group != ['*'] else order

        qry = _build_query(tname, cols, order)
        qrows = self.conn.cursor().execute(qry)
        columns = [c[0] for c in qrows.description]
        # there can't be duplicates in column names
        if len(columns) != len(set(columns)):
            raise ValueError('Duplicates in columns names')

        if where:
            def gen():
                for r in qrows:
                    r0 = _build_row(r, columns)
                    try:
                        if where(r0):
                            yield r0
                    except Exception:
                        pass
            rows = gen()
        else:
            rows = (_build_row(r, columns) for r in qrows)

        if overlap:
            size, step = (overlap, 1) if isnum(overlap) else overlap
            if group:
                gby = groupby(rows, _build_keyfn(group))
                grows = (Rows(rs) for _, rs in gby)
                yield from (Rows(chain(*xs))
                            for xs in _overlap(grows, size, step))
            else:
                yield from (Rows(rs) for rs in _overlap(rows, size, step))
        elif group:
            gby = groupby(rows, _build_keyfn(group))
            yield from (Rows(rs) for _, rs in gby)
        else:
            yield from rows

    def insert(self, rs, name):
        """Insert Rows or sequence of Row(s)

        Args:
            |  rs(Rows, or sequence of Row(s))
            |  name(str): table name
        """
        self.drop(name)
        rs = iter(rs)

        try:
            r0, rs = _peek_first(rs)
        except StopIteration:
            print(f'No Rows to Create: {name}')
            return

        cols = r0.columns
        n = len(cols)

        self._cursor.execute(_create_statement(name, cols))
        istmt = _insert_statement(name, n)
        self._cursor.executemany(istmt, (r.values for r in rs if isinstance(r, Row)))

    def load(self, filename, name=None, encoding='utf-8', fn=None):
        """Read data file and save it on database

        Args:
            |  filename(str): .csv, .xlsx, .sas7bdat
            |  name(str): table name
            |  encoding(str): file encoding
            |  fn(Row -> Row): Row transformer
        """
        if isinstance(filename, str):
            fname, ext = os.path.splitext(filename)

            if ext == '.csv':
                seq = _read_csv(filename, encoding)
            elif ext == '.xlsx':
                seq = _read_excel(filename)
            elif ext == '.sas7bdat':
                seq = _read_sas(filename)
            else:
                raise ValueError('Unknown file extension', ext)
            name = name or fname
        else:
            seq = filename

        if name in self.get_tables():
            return
        if fn:
            seq = (fn(r) for r in seq)

        self.insert(seq, name)


    def tocsv(self, tname, outfile=None, cols=None,
               where=None, order=None, encoding='utf-8'):
        """Table to csv file

        Args:
            |  tname(str): table name
            |  outfile(str): output file(csv) name
            |  cols(str or list of str)
            |  where(str)
            |  order(str)
            |  encoding(str)
        """
        seq = self.fetch(tname, cols=cols, where=where, order=order)
        r0, rs = _peek_first(seq)
        columns = _listify(cols) if cols else r0.columns
        filename = outfile or tname + '.csv'
        with open(os.path.join(WORKSPACE, filename), 'w', newline='',
                  encoding=encoding) as f:
            w = csv.writer(f, delimiter=',')
            w.writerow(columns)
            for r in rs:
                w.writerow(r.values)

    def get_tables(self):
        """List of table names in database
        """
        query = self._cursor.execute("""
        select * from sqlite_master
        where type='table'
        """)
        tables = [row[1] for row in query]
        return tables

    def drop(self, tables):
        """Drops tables if exist

        Args:
            tables(str, list of str)
        """
        tables = _listify(tables)
        for table in tables:
            # you can't use '?' for table name
            # '?' is for data insertion
            self._cursor.execute(f'drop table if exists {table}')

    # may or may not be deprecated
    def rename(self, old, new):
        """Rename a table from old to new
        """
        if old in self.get_tables():
            self._cursor.execute(f'drop table if exists { new }')
            self._cursor.execute(f'alter table { old } rename to { new }')

    # name must be specified
    def join(self, *tinfos, name=None):
        # rewrite tinfos if there's missing matching columns
        mcols0 = tinfos[0][2]

        temp_tables = []
        tinfos1 = []
        for tname, cols, mcols in tinfos:
            if hasattr(mcols, '__call__'):
                newtable = tname + '_' + _random_string(10)
                newcols = [newtable + '.' + c for c in _listify(cols)]
                newmcols = []

                def gen():
                    for r in self.fetch(tname):
                        try:
                            vals = mcols(r)
                            if not newmcols:
                                for v in vals:
                                    if (v == 0 or v):
                                        nc = 'col' + _random_string(10)
                                        newmcols.append(nc)
                                    else:
                                        newmcols.append('')
                            for c, v in zip(newmcols, vals):
                                r[c] = v
                            yield r
                        except Exception:
                            pass
                self.insert(gen(), newtable)
                tinfos1.append([newtable, newcols, newmcols])
                temp_tables.append(newtable)
            else:
                newcols = [tname + '.' + c for c in _listify(cols)]
                tinfos1.append([tname, newcols, _listify(mcols)])

        tname0, _, mcols0 = tinfos1[0]
        join_clauses = []
        for tname1, _, mcols1 in tinfos1[1:]:
            eqs = []
            for c0, c1 in zip(mcols0, mcols1):
                if c1:
                    eqs.append(f'{tname0}.{c0} = {tname1}.{c1}')
            join_clauses.append(f" left join {tname1} on {' and '.join(eqs)} ")
        jcs = ' '.join(join_clauses)

        allcols = []
        for i, (t1, cols, _) in enumerate(tinfos1):
            for c in cols:
                if c.endswith('.*'):
                    t0 = tinfos[i][0]
                    allcols += [t1 + '.' + c1 for c1 in self._cols(f'select * from {t0}')]
                else:
                    allcols.append(c)

        query = f"select {','.join(allcols)} from {tname0} {jcs}"
        self.create(query, name)
        self.drop(temp_tables)

    def create(self, query, name=None):
        """Create new table from query(select statement)
        Args:
            |  query(str)
            |  name(str): new table name, original table from the query
            |             if not exists
        """
        temp_name = 'table_' + _random_string()
        idx = query.lower().split().index('from')
        tname = query.split()[idx + 1]
        name = name or tname
        try:
            self._cursor.execute(_create_statement(temp_name, self._cols(query)))
            self._cursor.execute(f'insert into {temp_name} {query}')
            self._cursor.execute(f'drop table if exists { name }')
            self._cursor.execute(f"alter table { temp_name } rename to { name }")
        finally:
            self._cursor.execute(f'drop table if exists { temp_name }')

    def pwork(self, fn, tname, args):
        n = len(args)
        rndstr = _random_string()
        tempdbs = ['temp' + rndstr + str(i) for i in range(n)]
        try:
            with connect(tempdbs[0]) as c:
                c.insert(self.fetch(tname), tname)
            for tempdb in tempdbs[1:]:
                shutil.copyfile(os.path.join(WORKSPACE, tempdbs[0]),
                                os.path.join(WORKSPACE, tempdb))

            with ProcessPoolExecutor(max_workers=psutil.cpu_count(logical=False)) as exe:
                exe.map(fn, tempdbs, args)

            with connect(tempdbs[0]) as c:
                tables = [t for t in c.get_tables() if t != tname]

            for table in tables:
                self._collect(table, tempdbs)
        finally:
            for tempdb in tempdbs:
                fname = os.path.join(WORKSPACE, tempdb)
                if os.path.isfile(fname):
                    os.remove(fname)

    def _collect(self, tname, dbnames):
        with connect(dbnames[0]) as c:
            cols = c._cols(f"select * from {tname}")

        self.drop(tname)
        self._cursor.execute(_create_statement(tname, cols))
        ismt = _insert_statement(tname, len(cols))
        for dbname in dbnames:
            with connect(dbname) as c:
                self._cursor.executemany(ismt,
                                         (r.values for r in c.fetch(tname)))

    def _cols(self, query):
        return [c[0] for c in self._cursor.execute(query).description]

    def _pkeys(self, tname):
        "Primary keys in order"
        pks = [r for r in self._cursor.execute(f'pragma table_info({tname})') if r[5]]
        return [r[1] for r in sorted(pks, key=lambda r: r[5])]


def _build_keyfn(key):
    " if key is a string return a key function "
    # if the key is already a function, just return it
    if hasattr(key, '__call__'):
        return key
    colnames = _listify(key)
    # special case
    if colnames == ['*']:
        return lambda r: 1 

    if len(colnames) == 1:
        col = colnames[0]
        return lambda r: r[col]
    else:
        return lambda r: [r[colname] for colname in colnames]

# primary keys are too much for non-experts
def _create_statement(name, colnames):
    """create table if not exists foo (...)

    Note:
        Every type is numeric.
        Table name and column names are all lower cased
    """
    # every col is numeric, this may not be so elegant but simple to handle.
    # If you want to change this, Think again
    schema = ', '.join([col + ' ' + 'numeric' for col in colnames])
    return "create table if not exists %s (%s)" % (name, schema)


def _insert_statement(name, ncol):
    """insert into foo values (?, ?, ?, ...)
    """
    qmarks = ', '.join(['?'] * ncol)
    return "insert into %s values (%s)" % (name, qmarks)


def _build_query(tname, cols=None, order=None):
    "Build select statement"
    cols = ', '.join(_listify(cols)) if cols else '*'
    order = 'order by ' + ', '.join(_listify(order)) if order else ''
    return f'select {cols} from {tname} {order}'


def _build_row(qr, cols):
    r = Row()
    for c, v in zip(cols, qr):
        r[c] = v
    return r


def _read_csv(filename, encoding='utf-8'):
    "Loads well-formed csv file, 1 header line and the rest is data "
    def is_empty_line(line):
        """Tests if a list of strings is empty for example ["", ""] or []
        """
        return [x for x in line if x.strip() != ""] == []

    with open(os.path.join(WORKSPACE, filename),
              encoding=encoding) as fin:
        first_line = fin.readline()[:-1]
        columns = _listify(first_line)
        ncol = len(columns)

        # reader = csv.reader(fin)
        # NULL byte error handling
        reader = csv.reader(x.replace('\0', '') for x in fin)
        for line_no, line in enumerate(reader, 2):
            if len(line) != ncol:
                if is_empty_line(line):
                    continue
                raise ValueError(
                    """%s at line %s column count not matched %s != %s: %s
                    """ % (filename, line_no, ncol, len(line), line))
            yield _build_row(line, columns)


def _read_sas(filename):
    filename = os.path.join(WORKSPACE, filename)
    with SAS7BDAT(filename) as f:
        reader = f.readlines()
        # lower case
        header = next(reader)
        for line in reader:
            yield _build_row(line, header)


# this could be more complex but should it be?
def _read_excel(filename):
    def read_df(df):
        cols = df.columns
        for i, r in df.iterrows():
            yield _build_row((str(r[c]) for c in cols), cols)

    filename = os.path.join(WORKSPACE, filename)
    # it's OK. Excel files are small
    df = pd.read_excel(filename)
    yield from read_df(df)


def readxl(fname, sheet_name=None, encoding='utf-8'):
    def conv(x):
        try:
            return locale.atoi(x)
        except ValueError:
            try:
                return locale.atof(x)
            except:
                return x

    fname = os.path.join(WORKSPACE, fname)
    if fname.endswith('.csv'):
        with open(fname, encoding=encoding) as fin:
            for rs in csv.reader(x.replace('\0', '') for x in fin):
                yield [conv(x) for x in rs]
    else:
        workbook = load_workbook(fname)
        if not sheet_name:
            sheet_name = workbook.sheetnames[0]
        for row in workbook[sheet_name].iter_rows():
            yield [c.value for c in row]


def rename(old, new):
    with connect(DBNAME) as c:
        c.rename(old, new)


def drop(tables):
    with connect(DBNAME) as c:
        c.drop(tables)

 
def tocsv(tname, outfile=None, cols=None, where=None, order=None, encoding='utf-8'):
    with connect(DBNAME) as c:
        c.tocsv(tname, outfile, cols, where, order, encoding)


def process(*jobs):
    jobs = [job for job in jobs if not isinstance(job, str)]

    def build_parallel(jobs):
        def keyfn(x):
            if isinstance(x, Map):
                return x.inputs + [x.output]
            else:
                return x

        result = []
        for _, gs in groupby(jobs, keyfn):
            gs = list(gs)
            if len(gs) == 1:
                result.append(gs[0])
            else:
                g0 = gs[0]
                fns = [g1.fn for g1 in gs]
                selects = [g1.select for g1 in gs]
                args = [g1.arg for g1 in gs]
                result.append(Parallel(g0.inputs[0], g0.output,
                                       fns, args, selects))
        return result

    def find_required_tables(jobs):
        tables = set()
        for job in jobs:
            for table in job.inputs:
                tables.add(table)
            tables.add(job.output)
        return tables

    # depth first search
    def dfs(data, path, paths=[]):
        datum = path[-1]
        if datum in data:
            for val in data[datum]:
                new_path = path + [val]
                paths = dfs(data, new_path, paths)
        else:
            paths += [path]
        return paths

    def build_graph(jobs):
        graph = {}
        for job in jobs:
            for ip in job.inputs:
                if graph.get(ip):
                    graph[ip].add(job.output)
                else:
                    graph[ip] = {job.output}
        for x in graph:
            graph[x] = list(graph[x])
        return graph

    def render_graph(graph):
        dot = Digraph()
        for k, v in graph.items():
            dot.node(k, k)
            for v1 in v:
                dot.edge(k, v1)
        dot.render(GRAPH_NAME)

    required_tables = find_required_tables(jobs)
    with connect(DBNAME) as c:
        def delete_after(missing_table, paths):
            for path in paths:
                if missing_table in path:
                    for x in path[path.index(missing_table):]:
                        c.drop(x)

        def get_missing_tables():
            existing_tables = c.get_tables()
            return [table for table in required_tables
                    if table not in existing_tables]

        def find_jobs_to_do(jobs):
            missing_tables = get_missing_tables()
            result = []
            for job in jobs:
                for table in (job.inputs + [job.output]):
                    if table in missing_tables:
                        result.append(job)
                        break
            return result

        def is_doable(job):
            missing_tables = get_missing_tables()
            return all(table not in missing_tables for table in job.inputs) \
                and job.output in missing_tables
        jobs = build_parallel(jobs)

        outputs = [job.output for job in jobs]
        dups = set(x for x in outputs if outputs.count(x) > 1)
        assert len(dups) == 0, f"Dups: {dups}"

        graph = build_graph(jobs)

        try:
            render_graph(graph)
        except Exception:
            pass

        starting_points = [job.output for job in jobs if isinstance(job, Load)]
        paths = []
        for sp in starting_points:
            paths += dfs(graph, [sp], [])

        for mt in get_missing_tables():
            delete_after(mt, paths)

        jobs_to_do = find_jobs_to_do(jobs)
        print(f"Starting Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f'To Do: {[j.output for j in jobs_to_do]}')
        while jobs_to_do:
            cnt = 0
            for i, job in enumerate(jobs_to_do):
                if is_doable(job):
                    job.run(c)
                    tm = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    print(f"Done: {job.output} at {tm}")
                    del jobs_to_do[i]
                    cnt += 1
            if cnt == 0:
                print(f'Failed to Do: {[j.output for j in jobs_to_do]}')
                break
        

def add(**kwargs):
    def fn(r):
        for k, v in kwargs.items():
            try:
                r[k] = v(r)
            except Exception:
                r[k] = ''
        return r
    return fn


# They don't know pkeys, Just ignore it 
class Load:
    def __init__(self, filename, name=None, encoding='utf-8', fn=None):
        fname, ext = None, None
        if isinstance(filename, str):
            fname, ext = os.path.splitext(filename)

        self.filename = filename
        self.encoding = encoding
        self.fn = fn

        self.output = name or fname
        assert self.output is not None, "Table name must be provided"
        self.inputs = []

    def run(self, conn):
        if isinstance(self.fn, dict):
            fn = add(**self.fn)
        else:
            fn = self.fn
        conn.load(self.filename, self.output,
                  encoding=self.encoding, fn=fn)


class Union:
    def __init__(self, tables, name=None):
        self.inputs = _listify(tables)
        self.output = name 
        assert self.output not in self.inputs, """
        Output table name is one of the input table names
        """

    def run(self, conn):
        def gen():
            for input in self.inputs:
                for r in conn.fetch(input):
                    yield r 
        conn.insert(gen(), self.output)


class Join:
    def __init__(self, *tinfos, name=None):
        self.tinfos = tinfos
        self.name = name

        self.output = name or tinfos[0][0]
        self.inputs = [tinfo[0] for tinfo in tinfos]
        assert self.output not in self.inputs, """
        Output table name is one of the input table names
        """

    def run(self, conn):
        conn.join(*self.tinfos, name=self.name)


# fn for pwork in paralell
def buildfn(dbfile, argset):
    fn, arg, select, input, output = argset
    with connect(dbfile) as c:
        c.insert(genfn(c, fn, arg, select, input), output)


def genfn(c, fn, arg, select, input):
    if arg:
        for rs in c.fetch(input, **select):
            try:
                val = fn(rs, arg)
                if isinstance(val, collections.Iterable) or isinstance(val, Rows):
                    yield from val 
                else:
                    yield val
            except Exception:
                pass
    else:
        if isinstance(fn, dict):
            fn = add(**fn)

        for rs in c.fetch(input, **select):
            try:
                val = fn(rs)
                if isinstance(val, collections.Iterable) or isinstance(val, Rows):
                    yield from val
                else:
                    yield val 
            except Exception:
                pass


# This is for parallel work
class Parallel:
    def __init__(self, input, output, fns, args, selects):
        self.fns = fns
        self.selects = selects
        self.args = args

        self.inputs = [input]
        self.output = output
        assert self.inputs[0] != self.output, """
        Input and output table name must not be equal
        """

    def run(self, conn):
        conn.pwork(buildfn, self.inputs[0],
                   list(zip(self.fns,
                            self.args,
                            self.selects,
                            repeat(self.inputs[0]),
                            repeat(self.output))))


class Map:
    def __init__(self, fn, input, **kwargs):
        self.fn = fn
        self.select = {}
        self.arg = None

        for k, v in kwargs.items():
            if k.upper() == "ARG":
                self.arg = v
            elif k.upper() == "NAME":
                self.output = v
            else:
                self.select[k] = v

        self.inputs = [input]
        assert self.inputs[0] != self.output, """
        Input and output table name must not be equal
        """

    def run(self, conn):
        conn.insert(genfn(conn, self.fn, self.arg, self.select, self.inputs[0]), self.output)
